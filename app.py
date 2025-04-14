from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime

app = Flask(__name__)

FIELDS = [
    "구분", "계약여부", "식별번호", "계약금액", "제품모델명", "품명", "모델명",
    "규격", "수량", "원산지", "구성종류", "제품원가", "원천제조사", "수익률", "비고"
]

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/download_excel", methods=["POST"])
def download_excel():
    data = request.form.to_dict(flat=False)
    df = pd.DataFrame({k: v for k, v in data.items() if k in FIELDS})
    memo = data.get("업로드메모", [""])[0]

    # 마스터시트 템플릿 불러오기
    template_path = "master_template_copy.xlsx"
    wb = load_workbook(template_path)
    ws = wb["영상감시 마스터시트"]

    # 웹 입력값: '원산지' + '/' + '원천제조사' → '원산지 / 제조사'로 결합
    if "원산지" in df.columns and "원천제조사" in df.columns:
        df["원산지 / 제조사"] = df["원산지"].fillna('') + " / " + df["원천제조사"].fillna('')
    if "비고" in df.columns:
        df["비고"] = df["비고"].fillna('')

    # 열 순서 재정렬
    target_cols = [cell.value for cell in ws[1]]
    df = df.reindex(columns=[col for col in target_cols if col in df.columns])

    # 붙여넣기 시작 위치
    start_row = ws.max_row + 1
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 메모 시트 추가
    memo_ws = wb.create_sheet("업로드메모")
    memo_ws["A1"] = "업로드 메모"
    memo_ws["B1"] = memo

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"마스터시트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)